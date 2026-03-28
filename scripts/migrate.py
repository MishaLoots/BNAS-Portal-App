"""
BNAS Portal — Data Migration Script
Reads BNAS Artist Paybook 2026 V2.xlsx and inserts into Supabase.

Usage:
  pip install openpyxl supabase
  SUPABASE_URL=... SUPABASE_KEY=... python3 migrate.py

Set SUPABASE_URL and SUPABASE_KEY to your project's service role key
(Settings > API > service_role — NOT the anon key).
"""

import os, sys
from datetime import datetime, date
from openpyxl import load_workbook
from supabase import create_client

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]   # service_role key
EXCEL_PATH   = os.environ.get("EXCEL_PATH", "BNAS Artist Paybook 2026.xlsx")

sb = create_client(SUPABASE_URL, SUPABASE_KEY)
wb = load_workbook(EXCEL_PATH, data_only=True)

def d(val):
    """Convert to float, default 0."""
    if val is None: return 0
    try: return float(val)
    except: return 0

def to_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Try DD-Mon-YY format e.g. "02-Feb-26"
    try: return datetime.strptime(s, "%d-%b-%y").strftime("%Y-%m-%d")
    except: pass
    return s[:10]

def dep_values(raw):
    """Convert Y/N/Pre/0-100 dep value → (dep_pct, dep_is_pre)."""
    if raw is None or raw == "": return None, False
    if str(raw).strip().lower() == "pre": return None, True
    if str(raw).strip().upper() == "Y":  return 100, False
    if str(raw).strip().upper() == "N":  return 0, False
    try: return float(raw), False
    except: return None, False


# ── Step 1: Insert artists ───────────────────────────────────
ARTISTS = [
    {
        "name": "Pedro Barbosa",
        "email": "fatgrooveproductions@gmail.com",
        "escrow_account": "Pedro Escrow 2",
        "default_comm": 0.20,
        "default_warchest": 0.20,
        "opening_balance": 0,   # set from sheet below
        "loan_opening": 0,
        "has_csr": False,
    },
    {
        "name": "Apple Gule",
        "email": "applegule@bnas.co.za",
        "escrow_account": "AG Escrow",
        "default_comm": 0.25,
        "default_warchest": 0.20,
        "opening_balance": 0,
        "loan_opening": 5067,
        "has_csr": True,
    },
    {
        "name": "Saxby Twins",
        "email": "bookingssaxby@gmail.com",
        "escrow_account": "SXB Escrow",
        "default_comm": 0.10,
        "default_warchest": 0.20,
        "opening_balance": 0,
        "loan_opening": 0,
        "has_csr": False,
    },
    {
        "name": "Anzo",
        "email": "SlwaneMusic@outlook.com",
        "escrow_account": "Anzo Escrow",
        "default_comm": 0.20,
        "default_warchest": 0.20,
        "opening_balance": 0,
        "loan_opening": 6200,
        "has_csr": False,
    },
]

SHEET_MAP = {
    "Pedro Barbosa": "Pedro Barbosa",
    "Apple Gule":    "Apple Gule",
    "Saxby Twins":   "Saxby Twins",
    "Anzo":          "Anzo",
}

artist_ids = {}

print("Inserting artists...")
for a in ARTISTS:
    # Get opening balance from sheet
    sheet_name = SHEET_MAP.get(a["name"])
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Opening balance is in D84 in V2
        ob = ws["D84"].value
        a["opening_balance"] = d(ob)

    res = sb.table("artists").insert(a).execute()
    artist_id = res.data[0]["id"]
    artist_ids[a["name"]] = artist_id
    print(f"  ✓ {a['name']} → {artist_id}")


# ── Step 2: Insert shows ─────────────────────────────────────
print("\nInserting shows...")
DS, DE = 10, 60

for artist_name, sheet_name in SHEET_MAP.items():
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ Sheet '{sheet_name}' not found, skipping")
        continue

    ws  = wb[sheet_name]
    aid = artist_ids[artist_name]
    count = 0

    for r in range(DS, DE + 1):
        gross = ws.cell(r, 4).value
        date_val = ws.cell(r, 1).value
        if artist_name == "Pedro Barbosa" and r <= 15:
            print(f"    DEBUG row {r}: date={repr(date_val)} gross={repr(gross)}")
        if not gross: continue
        # Skip non-date rows (e.g. "TOTALS" summary rows)
        if isinstance(date_val, str) and date_val.strip().upper() == "TOTALS": continue

        dep_raw = ws.cell(r, 19 if artist_name in ("Saxby Twins","Anzo") else 19).value
        # Column mapping (matches V2 structure):
        # 1=Date 2=Event 3=Type 4=Gross 5=PayType 6=Comm%
        # 7=Comm(R) 8=Sound 9=Mus1 10=Mus2 11=Mus3 12=Other
        # 13=TotalBand 14=Subtotal 15=WC% 16=WC(R) 17=Nett
        # 18=Batch# 19=Status 20=Dep 21=Check
        dep_raw = ws.cell(r, 20).value
        status  = ws.cell(r, 19).value
        dep_pct, dep_is_pre = dep_values(dep_raw)

        show = {
            "artist_id":    aid,
            "show_date":    to_date(ws.cell(r, 1).value),
            "event":        ws.cell(r, 2).value,
            "show_type":    ws.cell(r, 3).value,
            "gross":        d(gross),
            "pay_type":     ws.cell(r, 5).value or "Escrow",
            "comm_pct":     d(ws.cell(r, 6).value),
            "sound":        d(ws.cell(r, 8).value),
            "mus1":         d(ws.cell(r, 9).value),
            "mus2":         d(ws.cell(r, 10).value),
            "mus3":         d(ws.cell(r, 11).value),
            "other_costs":  d(ws.cell(r, 12).value),
            "warchest_pct": d(ws.cell(r, 15).value),
            "batch_num":    str(ws.cell(r, 18).value) if ws.cell(r, 18).value else None,
            "status":       str(status) if status else None,
            "dep_pct":      dep_pct,
            "dep_is_pre":   dep_is_pre,
            "notes":        ws.cell(r, 21).value if ws.cell(r, 21).value else None,
        }
        sb.table("shows").insert(show).execute()
        count += 1

    print(f"  ✓ {artist_name}: {count} shows")


# ── Step 3: Insert transfers ─────────────────────────────────
print("\nInserting transfers...")

# Pedro transfers (from V2 data)
pedro_transfers = [
    ("2026-02-02", "Batch 1 — Jan 25-30",                       "Batch Payout",    12160),
    ("2026-02-09", "Batch 2 — 2-7 Feb",                         "Batch Payout",    35150),
    ("2026-02-16", "Batch 3 — 7-14 Feb",                        "Batch Payout",    23864),
    ("2026-02-23", "Batch 4 — 20-22 Feb",                       "Batch Payout",    6080),
    ("2026-03-02", "Batch 5 — 27-28 Feb (incl TN error +R5k)",  "Batch Payout",    29553),
    ("2026-03-02", "Warchest Distribution (Batch 5)",            "Warchest Dist.",  7000),
    ("2026-03-09", "Batch 6 — 6-7 Mar",                         "Batch Payout",    1500),
    ("2026-03-16", "Batch 7 — 12-15 Mar",                       "Batch Payout",    27500),
    ("2026-03-19", "Warchest Advance",                           "Warchest Dist.",  10000),
    ("2026-03-23", "Batch 8 Payout",                             "Batch Payout",    9473),
]

saxby_transfers = [
    ("2026-03-15", "Batch 1 — Mar payout", "Batch Payout", 7000),
]

anzo_transfers = [
    ("2026-03-19", "Batch 1 — Mar 2026", "Batch Payout", 5000),
]

apple_transfers = [
    ("2026-01-25", "Batch 1 — Jan payout", "Batch Payout", 25000),
]

for artist_name, transfers in [
    ("Pedro Barbosa", pedro_transfers),
    ("Apple Gule",    apple_transfers),
    ("Saxby Twins",   saxby_transfers),
    ("Anzo",          anzo_transfers),
]:
    aid = artist_ids[artist_name]
    for dt, desc, ttype, amt in transfers:
        sb.table("transfers").insert({
            "artist_id": aid, "transfer_date": dt,
            "description": desc, "transfer_type": ttype, "amount": amt,
        }).execute()
    print(f"  ✓ {artist_name}: {len(transfers)} transfers")


# ── Step 4: Insert payouts ───────────────────────────────────
print("\nInserting payouts...")

payouts_data = {
    "Pedro Barbosa": [("2026-03-15", "Batch 1-7", 77712.04, "")],
    "Apple Gule":    [("2026-01-25", "1",          2042,     "")],
    "Saxby Twins":   [("2026-03-15", "1",          6300,     "")],
    "Anzo":          [("2026-03-19", "Batch 1",    4000,     "")],
}

for artist_name, payouts in payouts_data.items():
    aid = artist_ids[artist_name]
    for dt, batch, amt, notes in payouts:
        sb.table("payouts").insert({
            "artist_id": aid, "payout_date": dt,
            "batch_ref": batch, "amount": amt, "notes": notes or None,
            "approved_by_artist": True,  # historical — pre-approved
            "approved_at": dt + "T00:00:00Z",
        }).execute()
    print(f"  ✓ {artist_name}: {len(payouts)} payouts")


# ── Step 5: Set admin profile ────────────────────────────────
print("\nSetting admin profile...")
print("  ⚠  Run setup_users.py first to create auth users,")
print("     then re-run this script's Step 5 manually if needed.")
print("  — or set is_admin=true in the profiles table via Supabase dashboard")

print("\n✅ Migration complete!")
print(f"   Artists: {len(ARTISTS)}")
